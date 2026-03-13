#!/usr/bin/env python3
"""生成 computeSalaryData 代码审查问题清单 Excel"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "computeSalaryData代码审查"

# 表头样式
header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 级别颜色
p0_fill = PatternFill(start_color="FF4D4D", end_color="FF4D4D", fill_type="solid")  # 红
p1_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # 橙
p2_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # 黄

level_fills = {"P0（严重）": p0_fill, "P1（重要）": p1_fill, "P2（一般）": p2_fill}
level_fonts = {
    "P0（严重）": Font(name="微软雅黑", bold=True, size=10, color="FFFFFF"),
    "P1（重要）": Font(name="微软雅黑", bold=True, size=10, color="000000"),
    "P2（一般）": Font(name="微软雅黑", bold=True, size=10, color="000000"),
}

thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap_align = Alignment(vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

headers = ["编号", "严重级别", "问题分类", "问题标题", "所在方法", "代码行号", "问题详细描述", "建议修复方向", "修改后效果"]
col_widths = [6, 12, 14, 30, 36, 10, 60, 45, 50]

for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

issues = [
    # P0
    [1, "P0（严重）", "冗余DB操作",
     "addAdditionalDeductionOptions 循环内逐条DB删除",
     "addAdditionalDeductionOptions",
     "4498",
     '每次调用会执行 lambdaUpdate().remove() 删除数据库记录，但该方法在 computeEmployeeSalary 的员工循环内调用。前面 getOrCreateRecordAndApplyAttendance 已做过一次全量删除，此处再按 code 删除属于冗余DB操作，与"内存计算、批量保存"的设计目标矛盾。',
     "将删除操作移至循环外批量执行，或在 getOrCreateRecordAndApplyAttendance 的全量删除时已覆盖这些 code，则此处删除可移除。",
     "消除员工循环内的逐条DB删除，减少数据库交互次数从 N 次降为 0 或 1 次，显著提升批量计薪性能，降低数据库负载。"],
    [2, "P0（严重）", "N+1查询",
     "shouldSkipTaxForRemark 循环内单条DB查询回退",
     "shouldSkipTaxForRemark",
     "1608",
     "当 employeeMap 中没有 isRemark 字段时，会调用 employeeService.getById(employeeId) 逐条查询数据库。如果 mapList 中大量员工缺少该字段，会产生 N+1 查询问题。",
     "在 doComputeSalaryData 开始时批量预加载 isRemark 数据到 Map，传入 ctx 上下文中。",
     "消除 N+1 查询回退路径，isRemark 数据统一从内存 Map 获取，查询次数从最坏 N 次降为 1 次批量查询，提升稳定性和性能。"],
    [3, "P0（严重）", "N+1查询",
     "fillAttendanceDataForEmployee 循环内查询假期抵扣和排班时长",
     "fillAttendanceDataForEmployee",
     "1803, 1856",
     "holidayDeductionService.queryHolidayDeduction 和 getWorkHours 在每个员工循环内单独查询数据库，未做批量预加载，是明显的循环内 N+1 查询。",
     "在 loadAttendanceSyncBatchData 中批量预加载假期抵扣和排班时长数据，放入 AttendanceSyncBatchData 统一管理。",
     "假期抵扣和排班时长数据一次性批量加载到内存，循环内零DB查询，整体计薪流程的DB交互减少 2N 次，大幅缩短计薪耗时。"],
    [4, "P0（严重）", "N+1查询",
     "fillAttendanceDataForEmployee 末尾冗余查询员工和部门信息",
     "fillAttendanceDataForEmployee",
     "2078-2081",
     "hrmEmployeeMapper.getEmployeeById(employeeId) 和 hrmDeptRepository.getAllByDeptId() 在每个员工循环内执行，但 map 中实际已包含部门信息，属于冗余查询。",
     "从 map 中直接获取已有的部门信息，或在循环外批量预加载。",
     "去除每个员工循环内 2 次冗余DB查询，直接复用内存中已有的员工和部门数据，减少 2N 次数据库调用，同时消除数据不一致风险。"],
    # P1
    [5, "P1（重要）", "硬编码",
     "硬编码员工ID散布多处",
     "getDeptTypeForEmployee, fillAttendanceDataForEmployee",
     "330-336, 1834, 1899, 2080",
     "如 \"1712718940198\"、\"1712718940181\"（董事长）、\"1481534121629855751\"（部门ID）等硬编码的员工ID和部门ID，业务规则绑定了具体数据，无法在数据库层面维护，也无法适应人员变动。",
     "将特殊员工/部门规则提取到数据库配置表或枚举配置中，通过查询获取。",
     "特殊员工/部门规则可通过后台配置管理，人员变动时无需修改代码和重新部署，降低维护成本，提升系统灵活性。"],
    [6, "P1（重要）", "硬编码",
     "硬编码全勤奖金额",
     "fillAttendanceDataForEmployee",
     "2083-2084",
     "全勤奖金额 \"500\"（总监）和 \"100\"（非总监）硬编码在代码中。",
     "从数据库配置表或薪资规则中读取全勤奖金额。",
     "全勤奖金额可在后台动态调整，无需改代码即可应对薪资政策变化，同时便于审计和追溯历史配置。"],
    [7, "P1（重要）", "精度风险",
     "double 浮点比较不安全",
     "fillAttendanceDataForEmployee",
     "1906",
     "empAttendanceSummary.getActualityDays() < normalDays 使用 double 类型直接做 < 比较，存在浮点精度问题，如 21.749999999 < 21.75 可能导致误判。",
     "改用 BigDecimal 进行比较，或使用容差值（epsilon）判断。",
     "消除浮点精度导致的全勤判断误判，确保出勤天数比较结果准确可靠，避免因精度问题导致员工全勤奖错发或漏发。"],
    [8, "P1（重要）", "一致性",
     "loadLastMonthTaxDataMap 跨年处理方式与其他代码不一致",
     "loadLastMonthTaxDataMap",
     "1361",
     "使用 int targetMonth = month - 1 手动处理跨年，而同文件其他地方使用 YearMonth.minusMonths(1)，风格不统一。",
     "统一使用 YearMonth.of(year, month).minusMonths(1) 处理跨月/跨年逻辑。",
     "跨年跨月逻辑统一用 YearMonth API 处理，消除手动计算导致的1月份跨年Bug风险（month-1=0），代码风格一致更易维护。"],
    [9, "P1（重要）", "数据流混乱",
     "removeFullAttendanceAndUnionFeeForMidMonthPromotion 被调用两次",
     "computeEmployeeSalary",
     "1218, 1256",
     "processMidMonthPromotionSalary 可能重新写入全勤奖/工会费，导致清理方法需调用两次。反映数据流向不清晰：生产数据的方法不应在清理之后再产生需要清理的数据。",
     "让 processMidMonthPromotionSalary 内部保证不再写入已清理的全勤奖/工会费，从根源解决问题。",
     "数据流单向清晰，全勤奖/工会费清理只需一次调用，消除重复清理的隐患，代码逻辑更易理解和维护，降低月中转正场景下薪资计算出错概率。"],
    # P2
    [10, "P2（一般）", "类型安全",
     "Map<String, Object> 作为员工数据载体缺乏类型安全",
     "doComputeSalaryData, computeEmployeeSalary 等",
     "996, 1059, 1714",
     "所有员工字段访问都通过字符串 key + 强制类型转换（如 (Integer)map.get(\"status\")），存在 ClassCastException 风险，且无法编译期检查字段名拼写错误。",
     "定义一个 EmployeeComputeContext 或类似的 POJO 类替代 Map<String, Object>。",
     "字段访问有编译期类型检查，IDE 可自动补全和重构，消除 ClassCastException 和字段名拼写错误风险，代码可读性显著提升。"],
    [11, "P2（一般）", "方法过长",
     "fillAttendanceDataForEmployee 方法超过380行",
     "fillAttendanceDataForEmployee",
     "1714-2100+",
     "该方法混合了变量初始化、考勤数据查询、全勤判断、迟到/早退/旷工/事假扣款计算、全勤奖计算等多个职责，且缩进风格不统一。",
     "按职责拆分为多个子方法：如 calculateLateMoney()、calculateFullAttendanceBonus() 等。",
     "单个方法职责清晰、长度可控（建议每个子方法不超过50行），便于独立测试和调试，新开发人员可快速定位特定业务逻辑。"],
    [12, "P2（一般）", "编码规范",
     "BigDecimal 创建方式不一致",
     "fillAttendanceDataForEmployee",
     "1738, 1740, 其他",
     "同一方法内混合使用 new BigDecimal(ZERO)、new BigDecimal(0)、BigDecimal.ZERO 三种方式创建零值。",
     "统一使用 BigDecimal.ZERO。",
     "代码风格统一，消除因使用 new BigDecimal(double) 构造器引入的潜在精度问题，减少不必要的对象创建，提升代码一致性。"],
    [13, "P2（一般）", "命名规范",
     "变量命名不符合 Java 规范",
     "fillAttendanceDataForEmployee",
     "1901, 1945, 1744, 1732, 1772",
     "ShuangXiuDays（应小写开头）、AbsenteeismDays（同上）、leaveOfsickDays（应为 leaveOfSickDays）、misscardCount（应为 missCardCount）、isProduce 三元表达式冗余。",
     "按 Java 驼峰命名规范统一修正。",
     "变量命名符合 Java 标准驼峰规范，IDE 不再产生命名警告，代码可读性提升，新成员无需猜测变量含义。"],
    [14, "P2（一般）", "冗余代码",
     "冗余的双重判断",
     "fillAttendanceDataForEmployee",
     "1911-1914",
     "外层已判断 \"4\".equals(entryStatus)，内层再次判断 entryStatus!=null && \"4\".equals(entryStatus)，完全冗余。",
     "删除内层冗余判断。",
     "消除逻辑重复，代码缩进层级减少一层，条件分支更清晰，避免后续维护时因修改一处而遗漏另一处导致逻辑不一致。"],
    [15, "P2（一般）", "语法瑕疵",
     "多余的分号",
     "fillAttendanceDataForEmployee",
     "1742",
     "BigDecimal leaveOfAbsenceDays = new BigDecimal(ZERO);; 行末有两个分号。",
     "删除多余的分号。",
     "消除语法层面的瑕疵，代码更整洁，通过静态代码检查工具（如 SonarQube）的扫描，不再产生不必要的警告。"],
    [16, "P2（一般）", "冗余代码",
     "未使用的局部变量",
     "fillAttendanceDataForEmployee",
     "1722, 1738, 1759, 1760",
     "post、deptName、leaveDays 声明后未使用；info 在方法开头获取但约300行后才使用。",
     "删除未使用的变量，info 移至实际使用处。",
     "减少无效变量占用，消除 IDE 和静态分析工具的警告，变量在使用处就近声明使作用域更清晰，便于理解数据依赖关系。"],
    [17, "P2（一般）", "异常处理",
     "getOrCreateRecordAndApplyAttendance 新建记录时缺少 null 检查",
     "getOrCreateRecordAndApplyAttendance",
     "1647-1652",
     "attendanceDataMap.get(jobNumber) 可能返回 null，后续直接 cv.get(2) 会抛出 NullPointerException。外层 try-catch 将异常吞掉变成通用的 ATTENDANCE_DATA_ERROR，不利于排查。",
     "在访问 cv 前增加 null 检查并抛出明确的异常信息，如 \"工号 xxx 无考勤数据\"。",
     "出错时日志直接定位到具体工号和原因，排查效率大幅提升，避免 NullPointerException 被通用 catch 吞掉导致问题难以追踪。"],
]

for row_idx, issue in enumerate(issues, 2):
    for col_idx, value in enumerate(issue, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.font = Font(name="微软雅黑", size=10)
        if col_idx in (1, 2, 6):
            cell.alignment = center_align
        else:
            cell.alignment = wrap_align
    # 级别列着色
    level_cell = ws.cell(row=row_idx, column=2)
    level = level_cell.value
    if level in level_fills:
        level_cell.fill = level_fills[level]
        level_cell.font = level_fonts[level]

# 冻结首行
ws.freeze_panes = "A2"
# 自动筛选
ws.auto_filter.ref = f"A1:I{len(issues)+1}"
# 行高
for row_idx in range(2, len(issues) + 2):
    ws.row_dimensions[row_idx].height = 60

# --- 汇总 sheet ---
ws2 = wb.create_sheet("汇总统计")
ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 10
ws2.column_dimensions["C"].width = 40

summary_headers = ["严重级别", "数量", "核心问题"]
for col_idx, h in enumerate(summary_headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

summary_data = [
    ["P0（严重）", 4, "循环内DB查询(N+1)、冗余DB删除"],
    ["P1（重要）", 5, "硬编码ID/金额、浮点比较、数据流混乱"],
    ["P2（一般）", 8, "类型安全、命名规范、冗余代码、方法过长"],
]
for row_idx, row_data in enumerate(summary_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.font = Font(name="微软雅黑", size=10)
        cell.alignment = center_align if col_idx <= 2 else wrap_align
    level_cell = ws2.cell(row=row_idx, column=1)
    level = level_cell.value
    if level in level_fills:
        level_cell.fill = level_fills[level]
        level_cell.font = level_fonts[level]

output_path = "/Users/jiangyongming/Project/hr/hainan/docs/computeSalaryData-code-review.xlsx"
wb.save(output_path)
print(f"已保存到: {output_path}")
